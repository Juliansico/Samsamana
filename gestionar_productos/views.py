from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from .models import Producto
from .forms import ProductoForm
from gestionar_categoria.models import Categoria
from gestionar_marca.models import Marca
from gestionar_presentacion.models import Presentacion
from gestionar_proveedor.models import Proveedor
import logging
import os
from reportlab.lib.utils import ImageReader
from django.views.decorators.cache import never_cache
from django.conf import settings
from io import BytesIO
from django.http import HttpResponse
from django.contrib.staticfiles import finders
from reportlab.lib.pagesizes import letter
from reportlab.lib.units import inch
from reportlab.pdfgen import canvas
from reportlab.platypus import Table, TableStyle
from reportlab.lib.colors import Color
from .models import Producto
from openpyxl.drawing.image import Image
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from io import BytesIO
from openpyxl.styles import Border, Side
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors
from reportlab.lib.styles import ParagraphStyle
# Create your views here.


def get_breadcrumbs(request):
    path = request.path.split('/')[1:]
    breadcrumbs = [{'title': 'Inicio', 'url': '/dashboard/'}]  # El inicio te lleva al dashboard
    url = ''
    for item in path:
        url += f'/{item}'
        breadcrumbs.append({'title': item.capitalize(), 'url': url})
    return breadcrumbs

@never_cache
def dashboard(request):
    return render(request, 'dashboard.html')
@never_cache
@login_required
def gestionar_productos(request):
    productos = Producto.objects.all()
    breadcrumbs = get_breadcrumbs(request)
    return render(request, 'gestionar_productos.html', {'productos': productos, 'breadcrumbs': breadcrumbs})

@never_cache
@login_required
def añadir_producto(request):
    if request.method == 'POST':
        form = ProductoForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Producto añadido con éxito.')
            return redirect('gestionar_productos')
    else:
        form = ProductoForm()
    breadcrumbs = get_breadcrumbs(request)
    return render(request, 'añadir_producto.html', {'form': form, 'breadcrumbs': breadcrumbs})


@never_cache
@login_required
def editar_producto(request, producto_id):
    producto = get_object_or_404(Producto, id=producto_id)
    if request.method == 'POST':
        form = ProductoForm(request.POST, instance=producto)
        if form.is_valid():
            form.save()
            messages.success(request, 'Producto actualizado con éxito.')
            return redirect('gestionar_productos')
    else:
        form = ProductoForm(instance=producto)
    breadcrumbs = get_breadcrumbs(request)
    return render(request, 'editar_producto.html', {'form': form, 'producto': producto, 'breadcrumbs': breadcrumbs})

@never_cache
@login_required
def filtrar_productos(request):
    estado_filtro = request.GET.get('estado', None)
    buscar = request.GET.get('buscar', '')
    precio_min = request.GET.get('precio_min', None)
    precio_max = request.GET.get('precio_max', None)
    categoria_id = request.GET.get('categoria', '')
    marca_id = request.GET.get('marca', '')
    presentacion_id = request.GET.get('presentacion', '')
    proveedor_id = request.GET.get('proveedor', '')

    productos = Producto.objects.all()

    if estado_filtro == 'activado':
        productos = productos.filter(estado=True)
    elif estado_filtro == 'inactivado':
        productos = productos.filter(estado=False)

    if buscar:
        productos = productos.filter(nombre__icontains=buscar)
    if precio_min:
        productos = productos.filter(precio__gte=precio_min)
    if precio_max:
        productos = productos.filter(precio__lte=precio_max)
    if categoria_id:
        productos = productos.filter(categoria_id=categoria_id)
    if marca_id:
        productos = productos.filter(marca_id=marca_id)
    if presentacion_id:
        productos = productos.filter(presentacion_id=presentacion_id)
    if proveedor_id:
        productos = productos.filter(proveedor_id=proveedor_id)

    context = {
        'productos': productos,
        'categorias': Categoria.objects.all(),
        'marcas': Marca.objects.all(),
        'presentaciones': Presentacion.objects.all(),
        'proveedor': Proveedor.objects.all(),
        'breadcrumbs': get_breadcrumbs(request)
    }

    return render(request, 'gestionar_productos.html', context)

@never_cache
@login_required
def activar_inactivar_producto(request, producto_id):
    producto = get_object_or_404(Producto, id=producto_id)
    producto.estado = not producto.estado
    producto.save()
    estado = "activado" if producto.estado else "inactivado"
    messages.success(request, f'Producto {estado} con éxito.')
    breadcrumbs = get_breadcrumbs(request)
    return redirect('gestionar_productos')


def reporte_productos_pdf(request):
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, topMargin=0.5*inch, bottomMargin=0.5*inch)
    width, height = letter

    # Estilos
    styles = getSampleStyleSheet()
    title_style = styles['Title']
    subtitle_style = styles['Heading2']
    normal_style = styles['Normal']

    # Crear contenido
    elements = []

    # Título y subtítulo
    elements.append(Paragraph("SAMSAMANA", title_style))
    elements.append(Paragraph("Reporte de Productos", subtitle_style))

    # Datos de la tabla
    data = [
        ["ID", "Nombre", "Marca", "Presentación", "Categoría", "Precio", "Unidad de medida", "Estado", "Proveedor"]
    ]

    productos = Producto.objects.all()
    for producto in productos:
        data.append([
            str(producto.id),
            producto.nombre,
            producto.marca.nombre,
            producto.presentacion.nombre,
            producto.categoria.nombre,
            str(producto.precio),
            producto.unidad_de_medida,
            'Activo' if producto.estado else 'Inactivo',
            producto.proveedor.nombre if producto.proveedor else 'Sin proveedor'
        ])

    # Estilos de la tabla
    table_style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#003366')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#f2f2f2')),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
    ])

    # Crear la tabla con anchos de columna ajustados
    col_widths = [0.5*inch, 1.5*inch, 1*inch, 1*inch, 1*inch, 0.8*inch, 0.8*inch, 0.8*inch, 1.5*inch]
    table = Table(data, colWidths=col_widths, repeatRows=1)
    table.setStyle(table_style)

    # Función para dividir el texto en dos líneas si es necesario
    def split_header(text, width, font="Helvetica-Bold", size=10):
        from reportlab.pdfbase.pdfmetrics import stringWidth
        if stringWidth(text, font, size) > width:
            words = text.split()
            result = []
            line = ""
            for word in words:
                if stringWidth(line + word, font, size) <= width:
                    line += word + " "
                else:
                    result.append(line.strip())
                    line = word + " "
            result.append(line.strip())
            return "\n".join(result)
        return text

    # Aplicar división de texto a los encabezados
    for i, header in enumerate(data[0]):
        p = Paragraph(split_header(header, col_widths[i] - 6), 
                ParagraphStyle('Header', parent=styles['Normal'], alignment=1, 
                                textColor=colors.white, fontName='Helvetica-Bold', fontSize=10))
        table._cellvalues[0][i] = p

    elements.append(table)

    # Generar PDF con marca de agua
    doc.build(elements)

    # Preparar respuesta
    buffer.seek(0)
    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="Reporte_productos.pdf"'
    
    return response



def reporte_productos_excel(request):
    # Crear buffer para el archivo Excel
    buffer = BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte de Productos"

    # Estilos de Excel
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="0066cc", end_color="0066cc", fill_type="solid")
    alignment_center = Alignment(horizontal="center", vertical="center")

    # Añadir logo en la primera fila
    logo_path = os.path.join(settings.STATIC_ROOT, 'img', 'Samsamanalogo1PNG.png')
    if os.path.exists(logo_path):
        img = Image(logo_path)
        img.width = 170  # Tamaño ajustado del logo
        img.height = 80
        ws.add_image(img, 'A1')  # Posicionar el logo en la celda A1

    # Título centrado en la fila 3
    ws.merge_cells('A3:I3')  # Combina celdas para el título en función del ancho de la tabla
    title_cell = ws['A3']
    title_cell.value = "TABLA PRODUCTOS - BALNEARIO SAMSAMANA"
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = alignment_center

    # Encabezados de la tabla en la fila 5
    headers = ["ID", "Nombre", "Marca", "Presentación", "Categoría", "Precio", "Unidad de medida", "Proveedor", "Estado"]
    for col_num, header in enumerate(headers, 1):  # Columnas desde A hasta I
        cell = ws.cell(row=5, column=col_num, value=header)  # Fila 5 para encabezados
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = alignment_center

    # Obtener datos de productos del modelo y agregar a la tabla
    productos = Producto.objects.all()
    for row_num, producto in enumerate(productos, start=6):  # Los datos empiezan desde la fila 6
        ws.cell(row=row_num, column=1, value=producto.id).alignment = alignment_center
        ws.cell(row=row_num, column=2, value=producto.nombre).alignment = alignment_center
        ws.cell(row=row_num, column=3, value=producto.marca.nombre).alignment = alignment_center
        ws.cell(row=row_num, column=4, value=producto.presentacion.nombre).alignment = alignment_center
        ws.cell(row=row_num, column=5, value=producto.categoria.nombre).alignment = alignment_center
        ws.cell(row=row_num, column=6, value=producto.precio).alignment = alignment_center
        ws.cell(row=row_num, column=7, value=producto.unidad_de_medida).alignment = alignment_center
        ws.cell(row=row_num, column=8, value=producto.proveedor.nombre).alignment = alignment_center  # Agregar el proveedor
        
        estado = 'Activo' if producto.estado else 'Inactivo'
        estado_font = Font(color="00FF00") if producto.estado else Font(color="FF0000")  # Verde para Activo, Rojo para Inactivo
        estado_cell = ws.cell(row=row_num, column=9, value=estado)
        estado_cell.alignment = alignment_center
        estado_cell.font = estado_font

    # Ajustar anchos de columna
    column_widths = [20, 20, 20, 20, 20, 15, 20, 20, 10]  # Anchos ajustados según los datos
    for col_num, width in enumerate(column_widths, start=1):  # Ajustar columnas desde A a I
        column_letter = get_column_letter(col_num)
        ws.column_dimensions[column_letter].width = width

    # Agregar bordes a la tabla
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    for row in ws[f'A5:I{ws.max_row}']:  # Agregar bordes desde la fila 5 hasta la última fila con datos
        for cell in row:
            cell.border = thin_border

    # Ajustar alturas de filas
    ws.row_dimensions[1].height = 50  # Altura para la fila del logo
    ws.row_dimensions[3].height = 30  # Altura para la fila del título
    ws.row_dimensions[5].height = 20  # Altura para la fila de encabezados

    # Guardar y devolver el archivo Excel
    wb.save(buffer)
    buffer.seek(0)
    response = HttpResponse(buffer.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="Reporte_productos_samsamana.xlsx"'
    return response
