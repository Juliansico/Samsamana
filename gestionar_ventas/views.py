from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import  get_user_model
from django.contrib.auth.decorators import login_required
from django.db.models import Sum
from openpyxl import Workbook
from reportlab.platypus import Image
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from django.contrib import messages
from .models import Venta, Producto, DetalleVenta
from .forms import VentaForm, DetalleVentaForm
from django.views.decorators.cache import never_cache
from reportlab.lib.styles import getSampleStyleSheet
from io import BytesIO
from django.http import HttpResponse
from django.conf import settings
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from reportlab.lib.utils import ImageReader
from reportlab.platypus import Table, TableStyle
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
import os
from openpyxl.drawing.image import Image
from django.forms import inlineformset_factory
from django.http import JsonResponse
from openpyxl.styles import Border, Side, Alignment, Font
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph

def get_breadcrumbs(request):
    path = request.path.split('/')[1:]
    breadcrumbs = [{'title': 'Inicio', 'url': '/dashboard/'}]
    url = ''
    for item in path:
        url += f'/{item}'
        breadcrumbs.append({'title': item.capitalize(), 'url': url})
    return breadcrumbs


@login_required
@never_cache
def dashboard(request):
    breadcrumbs = get_breadcrumbs(request)
    return render(request, 'dashboard.html', {'breadcrumbs': breadcrumbs})



@login_required
@never_cache
def crear_venta(request):
    DetalleVentaFormSet = inlineformset_factory(
        Venta, 
        DetalleVenta, 
        form=DetalleVentaForm, 
        extra=1,  
        can_delete=True
    )
    
    if request.method == 'POST':
        form = VentaForm(request.POST)
        formset = DetalleVentaFormSet(request.POST)
        
        if form.is_valid() and formset.is_valid():
            venta = form.save(commit=False)  # No guardamos aún la venta
            formset.instance = venta  # Asociar la venta a los detalles

            # Comprobación de que los detalles no están vacíos
            detalles_validos = any([detalle_form.cleaned_data for detalle_form in formset if detalle_form.cleaned_data])

            if detalles_validos:
                venta.save()
                formset.save()

                # Reducir el stock de los productos
                for detalle in venta.detalles.all():
                    producto = detalle.producto
                    if producto.stock < detalle.cantidad:
                        formset.add_error(None, f"Stock insuficiente para el producto {producto.nombre}")
                        return render(request, 'crear_venta.html', {'form': form, 'formset': formset, 'breadcrumbs': get_breadcrumbs(request)})

                    producto.stock -= detalle.cantidad  # Reducir stock
                    producto.save()

                # Calcular y guardar el valor total
                valor_total = sum(detalle.subtotal for detalle in venta.detalles.all())
                venta.valor_total = valor_total
                venta.save()
                
                messages.success(request, 'Venta creada con éxito.')
                return redirect('gestionar_ventas')
            else:
                messages.error(request, 'Debe agregar al menos un producto.')
        else:
            messages.error(request, 'Por favor llene todos los campos del formulario')
    else:
        form = VentaForm()
        formset = DetalleVentaFormSet()
    
    breadcrumbs = get_breadcrumbs(request)
    return render(request, 'crear_venta.html', {'form': form, 'formset': formset, 'breadcrumbs': breadcrumbs})

@login_required
@never_cache
def gestionar_ventas(request):
    ventas = Venta.objects.all()
    breadcrumbs = get_breadcrumbs(request)
    return render(request, 'gestionar_ventas.html', {'ventas': ventas, 'breadcrumbs': breadcrumbs})

@login_required
@never_cache
def detalle_venta(request, id):
    venta = get_object_or_404(Venta, id=id)
    breadcrumbs = get_breadcrumbs(request)
    if request.GET.get('format') == 'pdf':
        return exportar_detalle_venta_pdf(request, id)
    elif request.GET.get('format') == 'excel':
        return exportar_detalle_venta_excel(request, id)
    return render(request, 'detalle_venta.html', {'venta': venta, 'breadcrumbs': breadcrumbs})

@login_required
@never_cache
def eliminar_venta(request, id):
    # Obtener la venta que se desea eliminar
    venta = get_object_or_404(Venta, id=id)
    
    if request.method == 'POST':
        # Obtener los detalles de la venta usando el campo correcto
        detalles = DetalleVenta.objects.filter(venta=venta)  
        
        # Iterar sobre los detalles de la venta y devolver stock
        for detalle in detalles:
            detalle.producto.stock += detalle.cantidad  # Aumentar el stock
            detalle.producto.save()  # Guardar el producto con el nuevo stock

        # Eliminar la venta
        venta.delete()
        messages.success(request, 'La venta ha sido eliminada exitosamente.')
        return redirect('gestionar_ventas')

    breadcrumbs = get_breadcrumbs(request)
    return render(request, 'eliminar_venta.html', {'venta': venta, 'breadcrumbs': breadcrumbs})

@never_cache
def obtener_precio_producto(request, producto_id):
    try:
        producto = Producto.objects.get(id=producto_id)
        data = {
            'precio': str(producto.precio)  # Asegúrate de convertir el precio a string si es necesario
        }
        return JsonResponse(data)
    except Producto.DoesNotExist:
        return JsonResponse({'error': 'Producto no encontrado'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

@never_cache
def obtener_stock_producto(request, producto_id):
    try:
        producto = Producto.objects.get(id=producto_id)
        data = {
            'stock': producto.stock  # Asegúrate de que producto.stock sea accesible
        }
        return JsonResponse(data)
    except Producto.DoesNotExist:
        return JsonResponse({'error': 'Producto no encontrado'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

def reporte_ventas_pdf(request):
    buffer = BytesIO()
    p = canvas.Canvas(buffer, pagesize=letter)
    width, height = letter
    margin = 50
    row_height = 20
    y_position = height - margin

    # Cargar y añadir la marca de agua
    watermark_path = os.path.join(settings.STATIC_ROOT, 'img', 'Samsamanalogo1PNG.png')
    if os.path.exists(watermark_path):
        try:
            p.saveState()
            p.setFillAlpha(0.1)  # Ajustar la transparencia (0.1 = 10% visible)
            img = ImageReader(watermark_path)
            iw, ih = img.getSize()
            aspect = ih / float(iw)
            p.drawImage(img, x=(width - iw * aspect) / 2, y=(height - ih * aspect) / 2, 
                        width=iw * aspect, height=ih * aspect, mask='auto')
            p.restoreState()
        except Exception as e:
            print("Error al agregar la marca de agua:", e)

    # Añadir títulos
    p.setFont("Helvetica-Bold", 24)
    title_text = "SAMSAMANA"
    title_width = p.stringWidth(title_text, "Helvetica-Bold", 24)
    p.drawString((width - title_width) / 2, y_position, title_text)
    y_position -= 30

    p.setFont("Helvetica", 18)
    subtitle_text = "Reporte de Ventas"
    subtitle_width = p.stringWidth(subtitle_text, "Helvetica", 18)
    p.drawString((width - subtitle_width) / 2, y_position, subtitle_text)
    y_position -= 50

    # Crear tabla con datos de ventas
    column_widths = [30, 100, 80, 100, 80, 90]
    headers = ["ID", "Producto", "Cantidad", "Fecha", "Total", "Estado"]
    data = [headers]

    ventas = Venta.objects.all()
    for venta in ventas:
        detalles = venta.detalles.all()
        productos = ", ".join([detalle.producto.nombre for detalle in detalles])
        cantidades = ", ".join([str(detalle.cantidad) for detalle in detalles])
        total_venta = sum([detalle.subtotal for detalle in detalles])

        data.append([
            str(venta.id),
            productos,
            cantidades,
            venta.fecha.strftime('%Y-%m-%d %H:%M:%S'),
            f"{total_venta:.2f}",
            'Activo' if venta.estado else 'Inactivo'
        ])

    table = Table(data, colWidths=column_widths, rowHeights=row_height)

    # Estilo de la tabla
    style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#003366')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONT', (0, 0), (-1, 0), 'Helvetica-Bold', 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#f2f2f2')),
        ('FONT', (0, 1), (-1, -1), 'Helvetica', 10),
        ('INNERGRID', (0, 0), (-1, -1), 0.5, colors.black),
        ('BOX', (0, 0), (-1, -1), 1, colors.black),
    ])
    table.setStyle(style)

    table_width, table_height = table.wrap(width, height)
    table_x = (width - table_width) / 2
    table_y = y_position - len(data) * row_height

    table.drawOn(p, table_x, table_y)

    p.showPage()
    p.save()
    buffer.seek(0)

    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="Reporte_ventas.pdf"'
    return response




def reporte_ventas_excel(request):
    # Crear buffer para el archivo Excel
    buffer = BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte de Ventas"

    # Estilos de Excel
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="0066cc", end_color="0066cc", fill_type="solid")
    alignment_center = Alignment(horizontal="center", vertical="center")

    # Añadir logo en la celda A1
    logo_path = os.path.join(settings.STATIC_ROOT, 'img', 'Samsamanalogo1PNG.png')
    if os.path.exists(logo_path):
        img = Image(logo_path)
        img.width = 170  # Tamaño ajustado del logo
        img.height = 80
        ws.add_image(img, 'A1')  # Posicionar el logo en la celda A1

    # Título centrado en la fila 2
    ws.merge_cells('A2:F2')  # Combina celdas para el título
    title_cell = ws['A2']
    title_cell.value = "TABLA VENTAS - BALNEARIO SAMSAMANA"
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = alignment_center

    # Encabezados de la tabla en la fila 4
    headers = ["ID", "Producto", "Cantidad", "Fecha", "Total", "Estado"]
    for col_num, header in enumerate(headers, 1):  # Columnas desde A hasta F
        cell = ws.cell(row=4, column=col_num, value=header)  # Fila 4 para encabezados
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = alignment_center

    # Obtener datos de ventas del modelo y agregar a la tabla
    ventas = Venta.objects.all()
    row_num = 5  # Los datos empiezan desde la fila 5
    for venta in ventas:
        detalles = venta.detalles.all()

        # Concatenar productos y cantidades si hay más de uno
        productos = ", ".join([detalle.producto.nombre for detalle in detalles])
        cantidades = ", ".join([str(detalle.cantidad) for detalle in detalles])
        total_venta = sum([detalle.subtotal for detalle in detalles])

        # Añadir fila con datos de la venta
        ws.cell(row=row_num, column=1, value=venta.id).alignment = alignment_center
        ws.cell(row=row_num, column=2, value=productos).alignment = alignment_center
        ws.cell(row=row_num, column=3, value=cantidades).alignment = alignment_center
        ws.cell(row=row_num, column=4, value=venta.fecha.strftime('%Y-%m-%d %H:%M:%S')).alignment = alignment_center
        ws.cell(row=row_num, column=5, value=f"{total_venta:.2f}").alignment = alignment_center
        estado = 'Activo' if venta.estado else 'Inactivo'
        estado_font = Font(color="00FF00") if venta.estado else Font(color="FF0000")  # Verde para Activo, Rojo para Inactivo
        estado_cell = ws.cell(row=row_num, column=6, value=estado)
        estado_cell.alignment = alignment_center
        estado_cell.font = estado_font

        row_num += 1

    # Ajustar anchos de columna
    column_widths = [30, 30, 15, 20, 10, 10]  # Anchos ajustados según los datos
    for col_num, width in enumerate(column_widths, start=1):  # Ajustar columnas desde A a F
        column_letter = get_column_letter(col_num)
        ws.column_dimensions[column_letter].width = width

    # Agregar bordes a la tabla
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    for row in ws[f'A4:F{row_num - 1}']:  # Agregar bordes desde la fila 4 hasta la última fila con datos
        for cell in row:
            cell.border = thin_border

    # Ajustar alturas de filas
    ws.row_dimensions[1].height = 80  # Altura para la fila del logo
    ws.row_dimensions[2].height = 30  # Altura para la fila del título
    ws.row_dimensions[4].height = 20  # Altura para la fila de encabezados

    # Guardar y devolver el archivo Excel
    wb.save(buffer)
    buffer.seek(0)
    response = HttpResponse(buffer.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="Reporte_ventas_samsamana.xlsx"'
    return response
def exportar_detalle_venta_pdf(request, id):
    venta = get_object_or_404(Venta, id=id)
    buffer = BytesIO()
    p = canvas.Canvas(buffer, pagesize=letter)
    width, height = letter
    margin = 50
    y_position = height - margin

    # Cargar la imagen de marca de agua
    watermark_path = os.path.join(settings.STATIC_ROOT, 'img', 'Samsamanalogo1PNG.png')
    if os.path.exists(watermark_path):
        try:
            p.saveState()
            p.setFillAlpha(0.1)  # Ajusta la transparencia al 10%
            img = ImageReader(watermark_path)
            img_width, img_height = img.getSize()
            aspect = img_height / float(img_width)
            p.drawImage(img, x=(width - img_width * aspect) / 2, y=(height - img_height * aspect) / 2, 
                        width=img_width * aspect, height=img_height * aspect, mask='auto', preserveAspectRatio=True)
            p.restoreState()
        except Exception as e:
            print("Error al agregar la marca de agua:", e)

    # Añadir títulos
    p.setFont("Helvetica-Bold", 24)
    title_text = "SAMSAMANA"
    title_width = p.stringWidth(title_text, "Helvetica-Bold", 24)
    p.drawString((width - title_width) / 2, y_position, title_text)
    y_position -= 30

    p.setFont("Helvetica", 18)
    subtitle_text = f"Detalle de Venta #{venta.id}"
    subtitle_width = p.stringWidth(subtitle_text, "Helvetica", 18)
    p.drawString((width - subtitle_width) / 2, y_position, subtitle_text)
    y_position -= 50

    # Información de la venta
    info_data = [
        ['Usuario', venta.usuario.username],
        ['Fecha', venta.fecha.strftime('%Y-%m-%d %H:%M:%S')],
        ['Valor Total', f"${venta.valor_total:.2f}"],
    ]

    # Añadir tabla de información de la venta
    column_widths = [100, 200]
    p.setFont("Helvetica-Bold", 12)
    for label, value in info_data:
        p.drawString(margin, y_position, label)
        p.drawString(margin + column_widths[0], y_position, value)
        y_position -= 20  # Espaciado entre filas

    # Detalles de la venta
    y_position -= 20  # Espacio antes de la tabla de detalles
    p.setFont("Helvetica-Bold", 12)
    p.drawString(margin, y_position, "Detalles de la Venta")
    y_position -= 20

    # Encabezados y datos de la tabla de detalles
    headers = ["Producto", "Cantidad", "Precio", "Subtotal"]
    table_data = [headers]

    for detalle in venta.detalles.all():
        table_data.append([
            detalle.producto.nombre,
            str(detalle.cantidad),
            f"${detalle.producto.precio:.2f}",
            f"${detalle.subtotal:.2f}"
        ])

    # Crear y estilizar la tabla
    table = Table(table_data, colWidths=[150, 100, 100, 100])
    
    style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#003366')),  # Encabezado azul oscuro
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),  # Texto blanco en encabezados
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),  # Centrar texto
        ('GRID', (0, 0), (-1, -1), 1, colors.black),  # Bordes negros
        ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#f2f2f2')),  # Fondo gris claro para datos
        ('FONT', (0, 0), (-1, 0), 'Helvetica-Bold', 12),  # Fuente en encabezados
        ('FONT', (0, 1), (-1, -1), 'Helvetica', 10),  # Fuente en datos
        ('INNERGRID', (0, 0), (-1, -1), 0.5, colors.black),  # Bordes internos más delgados
        ('BOX', (0, 0), (-1, -1), 1, colors.black),  # Borde externo
    ])
    
    table.setStyle(style)

    # Ajustar la posición de la tabla
    table_width, table_height = table.wrap(width - 2 * margin, height)
    table_x = (width - table_width) / 2  # Centrar la tabla
    table_y = y_position - table_height - 20  # Espaciado adicional

    table.drawOn(p, table_x, table_y)

    # Finalizar el PDF
    p.showPage()
    p.save()
    buffer.seek(0)

    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename=detalle_venta_{venta.id}.pdf'
    return response
def exportar_detalle_venta_excel(request, id):
    venta = get_object_or_404(Venta, id=id)
    buffer = BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = f"Detalle Venta {venta.id}"

    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="0066cc", end_color="0066cc", fill_type="solid")
    alignment_center = Alignment(horizontal="center", vertical="center")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # Agregar logo
    logo_path = os.path.join(settings.STATIC_ROOT, 'img', 'Samsamanalogo1PNG.png')
    if os.path.exists(logo_path):
        img = Image(logo_path)
        img.width = 170
        img.height = 80
        ws.add_image(img, 'A1')

    # Título
    ws.merge_cells('A2:F2')
    title_cell = ws['A2']
    title_cell.value = f"DETALLE DE VENTA #{venta.id} - BALNEARIO SAMSAMANA"
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = alignment_center

    # Información de la venta
    info_data = [
        ['Usuario', venta.usuario.username],
        ['Fecha', venta.fecha.strftime('%Y-%m-%d %H:%M:%S')],
        ['Valor Total', f"${venta.valor_total:.2f}"],
    ]
    for row, (label, value) in enumerate(info_data, start=4):
        ws.cell(row=row, column=1, value=label).alignment = alignment_center
        ws.cell(row=row, column=2, value=value).alignment = alignment_center

    # Encabezados de los detalles
    headers = ["Producto", "Cantidad", "Precio", "Subtotal"]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=7, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = alignment_center

    # Detalles de la venta
    row_num = 8
    for detalle in venta.detalles.all():
        ws.cell(row=row_num, column=1, value=detalle.producto.nombre).alignment = alignment_center
        ws.cell(row=row_num, column=2, value=detalle.cantidad).alignment = alignment_center
        ws.cell(row=row_num, column=3, value=f"${detalle.producto.precio:.2f}").alignment = alignment_center
        ws.cell(row=row_num, column=4, value=f"${detalle.subtotal:.2f}").alignment = alignment_center
        row_num += 1

    # Aplicar bordes y centrar contenido
    for row in ws[f'A4:F{row_num - 1}']:
        for cell in row:
            cell.border = thin_border
            cell.alignment = alignment_center

    # Ajustar anchos de columna
    column_widths = [30, 15, 15, 15]
    for col_num, width in enumerate(column_widths, start=1):
        ws.column_dimensions[get_column_letter(col_num)].width = width

    # Ajustar alturas de filas
    ws.row_dimensions[1].height = 80  # Altura para la fila del logo
    ws.row_dimensions[2].height = 30  # Altura para la fila del título
    ws.row_dimensions[7].height = 20  # Altura para la fila de encabezados

    wb.save(buffer)
    buffer.seek(0)
    
    response = HttpResponse(buffer.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=detalle_venta_{venta.id}.xlsx'
    return response