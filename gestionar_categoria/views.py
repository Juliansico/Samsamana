from django.shortcuts import render, redirect, get_object_or_404
from io import BytesIO
from django.http import HttpResponse
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from reportlab.lib import colors
from reportlab.platypus import Table, TableStyle
from reportlab.lib.utils import ImageReader
import os
from openpyxl.drawing.image import Image
from django.conf import settings
from django.views.decorators.cache import never_cache
from django.contrib.auth import authenticate, login, logout, get_user_model
from django.contrib.auth.decorators import login_required
from django.views.decorators.http import require_http_methods
from django.contrib import messages
from .models import  Categoria
from .forms import  CategoriaForm
from openpyxl.utils import get_column_letter
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.lib import colors
from openpyxl.styles import Border, Side
# Create your views here.

def get_breadcrumbs(request):
    path = request.path.split('/')[1:]
    breadcrumbs = [{'title': 'Inicio', 'url': '/dashboard/'}]  # El inicio te lleva al dashboard
    url = ''
    for item in path:
        url += f'/{item}'
        breadcrumbs.append({'title': item.capitalize(), 'url': url})
    return breadcrumbs

@never_cache
def dashboard(request):
    return render(request, 'dashboard.html')
@never_cache
@login_required
def gestionar_categoria(request):
    categorias = Categoria.objects.all()
    breadcrumbs = get_breadcrumbs(request)
    return render(request, 'gestionar_categoria.html', {'categorias': categorias, 'breadcrumbs': breadcrumbs})


@never_cache
@login_required
def añadir_categoria(request):
    if request.method == 'POST':
        form = CategoriaForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Categoría añadida con éxito.')
            return redirect('gestionar_categoria')
    else:
        form = CategoriaForm()
    breadcrumbs = get_breadcrumbs(request)
    return render(request, 'añadir_categoria.html', {'form': form, 'breadcrumbs': breadcrumbs})


@never_cache
@login_required
def editar_categoria(request, categoria_id):
    categoria = get_object_or_404(Categoria, id=categoria_id)
    if request.method == 'POST':
        form = CategoriaForm(request.POST, instance=categoria)
        if form.is_valid():
            form.save()
            messages.success(request, 'Categoría actualizada con éxito.')
            return redirect('gestionar_categoria')
    else:
        form = CategoriaForm(instance=categoria)
    breadcrumbs = get_breadcrumbs(request)
    return render(request, 'editar_categoria.html', {'form': form, 'categoria': categoria, 'breadcrumbs': breadcrumbs})

@never_cache
@login_required
def activar_inactivar_categoria(request, categoria_id):
    categoria = get_object_or_404(Categoria, id=categoria_id)
    categoria.estado = not categoria.estado
    categoria.save()
    estado = "activada" if categoria.estado else "inactivada"
    messages.success(request, f'Categoría {estado} con éxito.')
    breadcrumbs = get_breadcrumbs(request)
    return redirect('gestionar_categoria', {'breadcrumbs': breadcrumbs})

@never_cache
@login_required
def filtrar_categorias(request):
    estado_filtro = request.GET.get('estado', None)
    buscar = request.GET.get('buscar', '')

    categorias = Categoria.objects.all()

    if estado_filtro == 'activado':
        categorias = categorias.filter(estado=True)
    elif estado_filtro == 'inactivado':
        categorias = categorias.filter(estado=False)

    if buscar:
        categorias = categorias.filter(nombre__icontains=buscar)

    context = {
        'categorias': categorias,
    }

    return render(request, 'gestionar_categoria.html', context)


def reporte_categorias_pdf(request):
    buffer = BytesIO()
    p = canvas.Canvas(buffer, pagesize=letter)
    width, height = letter
    margin = 50
    row_height = 20
    y_position = height - margin

    # Cargar la imagen de marca de agua
    watermark_path = os.path.join(settings.STATIC_ROOT, 'img', 'Samsamanalogo1PNG.png')
    if os.path.exists(watermark_path):
        try:
            p.saveState()
            p.setFillAlpha(0.1)  # Cambiar la opacidad para hacer la marca de agua más sutil
            img = ImageReader(watermark_path)
            img_width, img_height = img.getSize()
            aspect = img_height / float(img_width)
            p.drawImage(img, x=(width - img_width*aspect) / 2, y=(height - img_height*aspect) / 2, width=img_width*aspect, height=img_height*aspect, mask='auto', preserveAspectRatio=True)
            p.restoreState()
        except Exception as e:
            print("Error al agregar la marca de agua:", e)

    # Añadir títulos
    p.setFont("Helvetica-Bold", 24)
    title_text = "SAMSAMANA"
    title_width = p.stringWidth(title_text, "Helvetica-Bold", 24)
    p.drawString((width - title_width) / 2, y_position, title_text)
    y_position -= 30

    p.setFont("Helvetica", 18)
    subtitle_text = "Reporte de Categorías"
    subtitle_width = p.stringWidth(subtitle_text, "Helvetica", 18)
    p.drawString((width - subtitle_width) / 2, y_position, subtitle_text)
    y_position -= 50

    # Crear tabla con datos de categorías
    headers = ["ID", "Nombre", "Estado"]
    data = [headers]

    categorias = Categoria.objects.all()
    for categoria in categorias:
        data.append([
            str(categoria.id),
            categoria.nombre,
            'Activo' if categoria.estado else 'Inactivo'
        ])

    column_widths = [50, 200, 100]
    table = Table(data, colWidths=column_widths, rowHeights=row_height)

    # Estilo de la tabla
    style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#003366')),  # Fondo oscuro para encabezados
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),  # Texto blanco en encabezados
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),  # Bordes negros
        ('FONT', (0, 0), (-1, 0), 'Helvetica-Bold', 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#f2f2f2')),  # Fondo gris claro para datos
        ('FONT', (0, 1), (-1, -1), 'Helvetica', 10),
        ('INNERGRID', (0, 0), (-1, -1), 0.5, colors.black),  # Bordes internos más delgados
        ('BOX', (0, 0), (-1, -1), 1, colors.black),  # Borde externo
    ])
    table.setStyle(style)

    # Ajustar la posición de la tabla
    table_width, table_height = table.wrap(width, height)
    table_x = (width - table_width) / 2
    table_y = y_position - len(data) * row_height

    table.drawOn(p, table_x, table_y)

    p.showPage()
    p.save()
    buffer.seek(0)

    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="Reporte_categorias.pdf"'
    return response

def reporte_categorias_excel(request):

    buffer = BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte de Categorías"


    header_font = Font(bold=True, color="FFFFFF")  
    header_fill = PatternFill(start_color="0066cc", end_color="0066cc", fill_type="solid") 
    alignment_center = Alignment(horizontal="center", vertical="center")  

    
    active_font = Font(color="00FF00")  
    inactive_font = Font(color="FF0000")  


    logo_path = os.path.join(settings.STATIC_ROOT, 'img', 'Samsamanalogo1PNG.png')
    if os.path.exists(logo_path):
        img = Image(logo_path)
        img.width = 120  
        img.height = 70
    
        ws.add_image(img, 'A1')
        ws.column_dimensions['A'].width = 20 
        ws.row_dimensions[1].height = 70 

    # Añadir título
    ws.merge_cells('A4:C4') 
    title_cell = ws['A4']
    title_cell.value = "TABLA CATEGORÍAS - BALNEARIO SAMSAMANA"
    title_cell.font = Font(bold=True, size=12) 
    title_cell.alignment = alignment_center 

    # Añadir encabezados
    headers = ["ID", "Nombre", "Estado"]
    for col_num, header in enumerate(headers, 1): 
        cell = ws.cell(row=6, column=col_num, value=header)
        cell.font = header_font 
        cell.fill = header_fill  
        cell.alignment = alignment_center 

    # Añadir datos de categorías
    categorias = Categoria.objects.all()
    for row_num, categoria in enumerate(categorias, 7): 
        ws.cell(row=row_num, column=1, value=categoria.id).alignment = alignment_center
        ws.cell(row=row_num, column=2, value=categoria.nombre).alignment = alignment_center
        estado_cell = ws.cell(row=row_num, column=3, value='Activo' if categoria.estado else 'Inactivo')
        estado_cell.alignment = alignment_center
        estado_cell.font = active_font if categoria.estado else inactive_font  


    column_widths = [20, 30, 15] 
    for col_num, width in enumerate(column_widths, 1): 
        column_letter = get_column_letter(col_num)
        ws.column_dimensions[column_letter].width = width


    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    for row in ws[f'A6:C{ws.max_row}']:
        for cell in row:
            cell.border = thin_border

    # Ajustar altura de las filas
    ws.row_dimensions[4].height = 25 
    ws.row_dimensions[6].height = 20 


    wb.save(buffer)
    buffer.seek(0)

    # Preparar la respuesta HTTP
    response = HttpResponse(buffer.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="Reporte_categorias_samsamana.xlsx"'
    return response
