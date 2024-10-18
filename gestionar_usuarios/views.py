from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import user_passes_test
from django.db.models import Q
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.contrib.auth.models import Group
from .models import Usuario
from reportlab.platypus import Image
import os
from .forms import UsuarioForm
from django.views.decorators.cache import never_cache
from io import BytesIO
from django.http import HttpResponse
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from reportlab.lib import colors
from reportlab.platypus import Table, TableStyle
from reportlab.lib.utils import ImageReader
from django.conf import settings
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from reportlab.lib import colors
from .models import Usuario
from openpyxl.drawing.image import Image



@never_cache
def dashboard(request):
    if not request.user.is__active_user():
        messages.error(request, "Tu cuenta está inactiva. Por favor, contacta al administrador.")
        return redirect('logout')
    return render(request, 'dashboard.html')

@never_cache
@login_required
@user_passes_test(lambda u: u.is_superuser)
def gestionar_usuarios(request):
    query = request.GET.get('q')
    if query:
        usuarios = Usuario.objects.filter(
            Q(usuario__icontains=query) |
            Q(apellido__icontains=query) |
            Q(documento__icontains=query)
        )
    else:
        usuarios = Usuario.objects.all()
    
    context = {
        'usuarios': usuarios,
    }
    return render(request, 'gestionar_usuarios.html', context)

@never_cache
@login_required
@user_passes_test(lambda u: u.is_superuser)
def añadir_usuario(request):
    if request.method == 'POST':
        form = UsuarioForm(request.POST)
        if form.is_valid():
            usuario = form.save()
            # Remove any admin-related permissions
            usuario.user_permissions.clear()
            messages.success(request, 'Usuario añadido con éxito.')
            return redirect('gestionar_usuarios')
    else:
        form = UsuarioForm()
    return render(request, 'añadir_usuario.html', {'form': form})
@never_cache
@login_required
@user_passes_test(lambda u: u.is_superuser)  # Solo superusuario puede acceder
def editar_usuario(request, usuario_id):
    usuario = get_object_or_404(Usuario, id=usuario_id)

    if request.method == 'POST':
        form = UsuarioForm(request.POST, instance=usuario)

        if form.is_valid():
            usuario = form.save(commit=False)

            # Solo cambiar la contraseña si se ha proporcionado una nueva
            if form.cleaned_data['password1']:  
                usuario.set_password(form.cleaned_data['password1'])

            # Asegurarse de que no sea superusuario
            usuario.is_superuser = False  
            usuario.is_staff = True  # Permitir acceso al admin

            usuario.save()

            # Asignación de grupo (rol de usuario) solo si se ha seleccionado un rol
            rol_usuario = form.cleaned_data.get('rol_usuario')
            if rol_usuario:
                try:
                    group = Group.objects.get(name=rol_usuario)
                    usuario.groups.clear()  # Limpiar grupos anteriores
                    usuario.groups.add(group)  # Asignar nuevo grupo
                except Group.DoesNotExist:
                    messages.error(request, f'El grupo "{rol_usuario}" no existe.')
                    return render(request, 'editar_usuario.html', {'form': form, 'usuario': usuario})
            
            # Mostrar mensaje de éxito si todo está bien
            messages.success(request, 'Usuario actualizado con éxito.')
            return redirect('gestionar_usuarios')

    else:
        form = UsuarioForm(instance=usuario)

    return render(request, 'editar_usuario.html', {'form': form, 'usuario': usuario})


@never_cache
@login_required
@user_passes_test(lambda u: u.is_superuser)
def activar_inactivar_usuario(request, usuario_id):
    usuario = get_object_or_404(Usuario, id=usuario_id)
    
    # Cambiar el estado del usuario
    usuario.estado = not usuario.estado
    usuario.save()  # Esto actualizará automáticamente `is_active` debido al método `save()`

    estado = "activado" if usuario.estado else "inactivado"  # Mostrar el estado correcto
    messages.success(request, f'Usuario {estado} con éxito.')
    return redirect('gestionar_usuarios')

@never_cache
@login_required
@user_passes_test(lambda u: u.is_superuser)
def filtrar_usuarios(request):
    usuarios = Usuario.objects.all()

    nombre = request.GET.get('nombre')
    documento = request.GET.get('documento')
    telefono = request.GET.get('telefono')
    estado = request.GET.get('estado')

    if nombre:
        usuarios = usuarios.filter(username__icontains=nombre)
    if documento:
        usuarios = usuarios.filter(documento__icontains=documento)
    if telefono:
        usuarios = usuarios.filter(telefono__icontains=telefono)
    if estado:
        if estado == 'activado':
            usuarios = usuarios.filter(estado=True)
        elif estado == 'inactivado':
            usuarios = usuarios.filter(estado=False)

    return render(request, 'gestionar_usuarios.html', {'usuarios': usuarios})

def reporte_usuarios_pdf(request):
    buffer = BytesIO()
    p = canvas.Canvas(buffer, pagesize=letter)
    width, height = letter
    margin = 50
    row_height = 20
    y_position = height - margin

    # Cargar la imagen de marca de agua
    watermark_path = os.path.join(settings.STATIC_ROOT, 'img', 'Samsamanalogo1PNG.png')
    if os.path.exists(watermark_path):
        try:
            p.saveState()
            p.setFillAlpha(0.1)  # Ajusta la transparencia al 10%
            img = ImageReader(watermark_path)
            img_width, img_height = img.getSize()
            aspect = img_height / float(img_width)
            p.drawImage(img, x=(width - img_width * aspect) / 2, y=(height - img_height * aspect) / 2, 
                        width=img_width * aspect, height=img_height * aspect, mask='auto', preserveAspectRatio=True)
            p.restoreState()
        except Exception as e:
            print("Error al agregar la marca de agua:", e)

    # Añadir títulos
    p.setFont("Helvetica-Bold", 24)
    title_text = "SAMSAMANA"
    title_width = p.stringWidth(title_text, "Helvetica-Bold", 24)
    p.drawString((width - title_width) / 2, y_position, title_text)
    y_position -= 30

    p.setFont("Helvetica", 18)
    subtitle_text = "Reporte de Usuarios"
    subtitle_width = p.stringWidth(subtitle_text, "Helvetica", 18)
    p.drawString((width - subtitle_width) / 2, y_position, subtitle_text)
    y_position -= 50

    # Crear tabla con datos de usuarios
    column_widths = [50, 150, 100, 200, 100]
    headers = ["ID", "Usuario", "Apellido", "Teléfono", "Estado"]
    data = [headers]

    usuarios = Usuario.objects.all()
    for usuario in usuarios:
        data.append([
            str(usuario.id),
            usuario.username,
            usuario.apellido,
            usuario.telefono,
            'Activo' if usuario.estado else 'Inactivo'
        ])

    table = Table(data, colWidths=column_widths, rowHeights=row_height)

    # Estilo de la tabla
    style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#003366')),  # Fondo oscuro para encabezados
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),  # Texto blanco en encabezados
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),  # Bordes negros
        ('FONT', (0, 0), (-1, 0), 'Helvetica-Bold', 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#f2f2f2')),  # Fondo gris claro para datos
        ('FONT', (0, 1), (-1, -1), 'Helvetica', 10),
        ('INNERGRID', (0, 0), (-1, -1), 0.5, colors.black),  # Bordes internos más delgados
        ('BOX', (0, 0), (-1, -1), 1, colors.black),  # Borde externo
    ])
    table.setStyle(style)

    # Ajustar la posición de la tabla
    table_width, table_height = table.wrap(width, height)
    table_x = (width - table_width) / 2
    table_y = y_position - len(data) * row_height

    table.drawOn(p, table_x, table_y)

    p.showPage()
    p.save()
    buffer.seek(0)

    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="Reporte_usuarios.pdf"'
    return response

def reporte_usuarios_excel(request):
    # Crear un archivo Excel en memoria
    buffer = BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte de Usuarios"

    # Definir estilos
    header_font = Font(bold=True, color="FFFFFF")  # Fuente blanca y en negrita para encabezados
    header_fill = PatternFill(start_color="0066cc", end_color="0066cc", fill_type="solid")  # Fondo azul oscuro para encabezados
    alignment_center = Alignment(horizontal="center", vertical="center")  # Alineación centrada

    # Colores para el estado
    active_font = Font(color="00FF00")  # Verde para Activo
    inactive_font = Font(color="FF0000")  # Rojo para Inactivo

    # Añadir logo
    logo_path = os.path.join(settings.STATIC_ROOT, 'img', 'Samsamanalogo1PNG.png')
    if os.path.exists(logo_path):
        img = Image(logo_path)
        img.width = 180  # Ajustar tamaño del logo
        img.height = 100
        ws.add_image(img, 'A1')  # Colocar el logo en la celda B2

        # Ajustar el ancho de las columnas para centrar el logo visualmente
        
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 15

    # Añadir título
    ws.merge_cells('B4:F4')  # Combina celdas para centrar el título
    title_cell = ws['B4']
    title_cell.value = "TABLA USUARIOS - BALNEARIO SAMSAMANA"
    title_cell.font = Font(bold=True, size=16)  # Fuente en negrita y tamaño 16 para el título
    title_cell.alignment = alignment_center  # Alineación centrada

    # Añadir encabezados
    headers = ["ID", "Usuario", "Apellido", "Teléfono", "Estado"]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=6, column=col_num, value=header)
        cell.font = header_font  # Aplicar fuente blanca en negrita
        cell.fill = header_fill  # Aplicar fondo azul oscuro
        cell.alignment = alignment_center  # Alineación centrada

    # Añadir datos de usuarios
    usuarios = Usuario.objects.all()
    for row_num, usuario in enumerate(usuarios, 7):  # Los datos empiezan desde la fila 7
        ws.cell(row=row_num, column=1, value=usuario.id).alignment = alignment_center
        ws.cell(row=row_num, column=2, value=usuario.username).alignment = alignment_center
        ws.cell(row=row_num, column=3, value=usuario.apellido).alignment = alignment_center
        ws.cell(row=row_num, column=4, value=usuario.telefono).alignment = alignment_center
        estado_cell = ws.cell(row=row_num, column=5, value='Activo' if usuario.estado else 'Inactivo')
        estado_cell.alignment = alignment_center
        estado_cell.font = active_font if usuario.estado else inactive_font  # Aplicar color verde o rojo según el estado

    # Ajustar el ancho de las columnas
    column_widths = [20, 20, 20, 20, 15]  # Ajustar los anchos según sea necesario
    for col_num, width in enumerate(column_widths, 1):
        column_letter = get_column_letter(col_num)
        ws.column_dimensions[column_letter].width = width
        

    # Añadir bordes a las celdas de la tabla
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    for row in ws[f'A6:E{ws.max_row}']:  # Aplicar bordes solo de A hasta E
        for cell in row:
            cell.border = thin_border

    # Ajustar altura de filas
    ws.row_dimensions[2].height = 40  # Altura de la fila del logo
    ws.row_dimensions[4].height = 25  # Altura de la fila del título
    ws.row_dimensions[6].height = 20  # Altura de la fila de los encabezados

    # Guardar y devolver
    wb.save(buffer)
    buffer.seek(0)
    response = HttpResponse(buffer.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="Reporte_usuarios.xlsx"'
    return response