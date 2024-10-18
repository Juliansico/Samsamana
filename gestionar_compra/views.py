from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from .models import  Compra, DetalleCompra, Producto, Proveedor
from .forms import  CompraForm, DetalleCompraForm
from gestionar_ventas.models import Venta
from reportlab.lib import colors
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image
from reportlab.lib.utils import ImageReader
from django.views.decorators.cache import never_cache
from io import BytesIO
from django.http import HttpResponse
from django.conf import settings
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from reportlab.lib.units import inch
from reportlab.platypus import Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import inch
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import os
from django.forms import inlineformset_factory
from django.http import JsonResponse
# Create your views here.
@never_cache
def dashboard(request):
    return render(request, 'dashboard.html')

@login_required
@never_cache
def crear_compra(request):
    DetalleCompraFormSet = inlineformset_factory(
        Compra, 
        DetalleCompra, 
        form=DetalleCompraForm, 
        extra=1,  
        can_delete=True
    )
    
    if request.method == 'POST':
        form = CompraForm(request.POST)
        formset = DetalleCompraFormSet(request.POST)
        
        if form.is_valid() and formset.is_valid():
            compra = form.save()
            formset.instance = compra 
            formset.save()

            # Calcular y guardar el valor total
            total_compra = sum(detalle.subtotal for detalle in compra.detalles.all())
            compra.total_compra = total_compra
            compra.save()
            
            return redirect('gestionar_compras')
        else:
            # Depuración de errores
            print(form.errors, formset.errors)  
    else:
        form = CompraForm()
        formset = DetalleCompraFormSet()
    
    return render(request, 'crear_compra.html', {'form': form, 'formset': formset})

@login_required
@never_cache
def gestionar_compras(request):
    compras = Compra.objects.all()
    return render(request, 'gestionar_compras.html', {'compras': compras})

@login_required
@never_cache
def detalle_compra(request, id):
    compra = get_object_or_404(Compra, id=id)
    return render(request, 'detalle_compra.html', {'compra': compra})

@login_required
@never_cache
def eliminar_compra(request, id):
    compra = get_object_or_404(Compra, id=id)
    
    if request.method == 'POST':
        # Obtener todos los productos de los detalles de la compra
        productos = compra.detalles.values_list('producto', flat=True)
        
        # Verificar si hay ventas relacionadas con esos productos
        if Venta.objects.filter(detalles__producto__in=productos).exists():
            messages.error(request, 'No se puede eliminar la compra porque ya se han realizado ventas relacionadas.')
            return redirect('gestionar_compras')
        
        # Actualizar stock de los productos
        for detalle in compra.detalles.all():
            detalle.producto.stock -= detalle.cantidad
            detalle.producto.save()

        # Eliminar la compra
        compra.delete()
        messages.success(request, 'La compra ha sido eliminada exitosamente y el stock ha sido actualizado.')
        return redirect('gestionar_compras')

    return render(request, 'eliminar_compra.html', {'compra': compra})





@never_cache
def obtener_precio_y_proveedor(request, producto_id):
    try:
        producto = Producto.objects.get(id=producto_id)
        data = {
            'precio': producto.precio,
            'proveedor': producto.proveedor.nombre,  # Asegúrate de que "proveedor" sea un campo relacionado con el producto
        }
        return JsonResponse(data)
    except Producto.DoesNotExist:
        return JsonResponse({'error': 'Producto no encontrado'}, status=404)
    
def reporte_compras_pdf(request):
    buffer = BytesIO()
    p = canvas.Canvas(buffer, pagesize=letter)
    width, height = letter
    margin = 50
    row_height = 20
    y_position = height - margin

    # Cargar y añadir la marca de agua
    watermark_path = os.path.join(settings.STATIC_ROOT, 'img', 'Samsamanalogo1PNG.png')
    if os.path.exists(watermark_path):
        try:
            p.saveState()
            p.setFillAlpha(0.1)  # Ajustar la transparencia (0.1 = 10% visible)
            img = ImageReader(watermark_path)
            iw, ih = img.getSize()
            aspect = ih / float(iw)
            # Centrar la imagen
            p.drawImage(img, x=(width - iw * aspect) / 2, y=(height - ih * aspect) / 2,
                        width=iw * aspect, height=ih * aspect, mask='auto')
            p.restoreState()
        except Exception as e:
            print("Error al agregar la marca de agua:", e)

    # Añadir títulos
    p.setFont("Helvetica-Bold", 30)  # Aumentar el tamaño de la fuente
    title_text = "SAMSAMANA"
    title_width = p.stringWidth(title_text, "Helvetica-Bold", 30)
    p.drawString((width - title_width) / 2, y_position, title_text)
    y_position -= 40  # Ajustar el espacio para el título

    p.setFont("Helvetica", 20)  # Aumentar el tamaño de la fuente para el subtítulo
    subtitle_text = "Reporte de Compras"
    subtitle_width = p.stringWidth(subtitle_text, "Helvetica", 20)
    p.drawString((width - subtitle_width) / 2, y_position, subtitle_text)
    y_position -= 50

    # Crear tabla con datos de compras
    column_widths = [30, 100, 80, 100, 80, 90]
    headers = ["ID", "Fecha Compra", "Total Compra", "Cantidad Producto", "Proveedor", "Estado"]
    data = [headers]

    compras = Compra.objects.all()
    for compra in compras:
        data.append([
            str(compra.id),
            compra.fecha.strftime('%Y-%m-%d %H:%M:%S'),
            f"{compra.total_compra:.2f}",
            str(compra.cantidad_productos),
            compra.usuario.username,
            'Activo' if compra.estado else 'Inactivo'
        ])

    table = Table(data, colWidths=column_widths, rowHeights=row_height)

    # Estilo de la tabla
    style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#003366')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#f2f2f2')),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('INNERGRID', (0, 0), (-1, -1), 0.5, colors.black),
        ('BOX', (0, 0), (-1, -1), 1, colors.black),
    ])
    table.setStyle(style)

    # Ajustar posición de la tabla
    table_width, table_height = table.wrap(width, height)
    table_x = (width - table_width) / 2
    table_y = y_position - len(data) * row_height

    table.drawOn(p, table_x, table_y)

    p.showPage()
    p.save()
    buffer.seek(0)

    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="Reporte_compras.pdf"'
    return response




from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from io import BytesIO
import os
from django.conf import settings

def reporte_compras_excel(request):
    buffer = BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte de Compras"

    # Estilos de Excel
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="0066cc", end_color="0066cc", fill_type="solid")
    alignment_center = Alignment(horizontal="center", vertical="center")

    # Añadir logo en la celda A1
    logo_path = os.path.join(settings.STATIC_ROOT, 'img', 'Samsamanalogo1PNG.png')
    if os.path.exists(logo_path):
        img = Image(logo_path)
        img.width = 170  # Tamaño ajustado del logo
        img.height = 80
        ws.add_image(img, 'A1')  # Posicionar el logo en la celda A1

    # Título centrado en la fila 2
    ws.merge_cells('A2:F2')  # Combina celdas para el título
    title_cell = ws['A2']
    title_cell.value = "TABLA COMPRAS - BALNEARIO SAMSAMANA"
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    # Encabezados de la tabla en la fila 4
    headers = ["ID", "Fecha Compra", "Total Compra", "Cantidad Producto", "Proveedor", "Estado"]
    for col_num, header in enumerate(headers, 1):  # Columnas desde A hasta F
        cell = ws.cell(row=4, column=col_num, value=header)  # Fila 4 para encabezados
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = alignment_center

    # Datos de compras
    compras = Compra.objects.all()
    row_num = 5  # Los datos empiezan desde la fila 5
    for compra in compras:
        ws.cell(row=row_num, column=1, value=compra.id).alignment = alignment_center
        ws.cell(row=row_num, column=2, value=compra.fecha.strftime('%Y-%m-%d %H:%M:%S')).alignment = alignment_center
        ws.cell(row=row_num, column=3, value=compra.total_compra).alignment = alignment_center
        ws.cell(row=row_num, column=4, value=compra.cantidad_productos).alignment = alignment_center
        ws.cell(row=row_num, column=5, value=compra.usuario.username).alignment = alignment_center
        estado_cell = ws.cell(row=row_num, column=6, value='Activo' if compra.estado else 'Inactivo')
        estado_cell.alignment = alignment_center

        row_num += 1

    # Ajustar anchos de columna
    column_widths = [10, 20, 15, 20, 30, 10]  # Anchos ajustados según los datos
    for col_num, width in enumerate(column_widths, start=1):  # Ajustar columnas desde A a F
        column_letter = get_column_letter(col_num)
        ws.column_dimensions[column_letter].width = width

    # Agregar bordes a la tabla
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    for row in ws[f'A4:F{row_num - 1}']:  # Agregar bordes desde la fila 4 hasta la última fila con datos
        for cell in row:
            cell.border = thin_border

    # Ajustar alturas de filas
    ws.row_dimensions[1].height = 80  # Altura para la fila del logo
    ws.row_dimensions[2].height = 30  # Altura para la fila del título
    ws.row_dimensions[4].height = 20  # Altura para la fila de encabezados

    # Guardar y devolver el archivo
    wb.save(buffer)
    buffer.seek(0)
    response = HttpResponse(buffer.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="Reporte_compras_samsamana.xlsx"'
    return response
def exportar_detalle_compra_pdf(request, id):
    compra = get_object_or_404(Compra, id=id)
    buffer = BytesIO()
    p = canvas.Canvas(buffer, pagesize=letter)
    width, height = letter
    margin = 50
    y_position = height - margin

    # Cargar la imagen de marca de agua
    watermark_path = os.path.join(settings.STATIC_ROOT, 'img', 'Samsamanalogo1PNG.png')
    if os.path.exists(watermark_path):
        try:
            p.saveState()
            p.setFillAlpha(0.1)  # Ajusta la transparencia al 10%
            img = ImageReader(watermark_path)
            img_width, img_height = img.getSize()
            aspect = img_height / float(img_width)
            p.drawImage(img, x=(width - img_width * aspect) / 2, y=(height - img_height * aspect) / 2,
                         width=img_width * aspect, height=img_height * aspect, mask='auto', preserveAspectRatio=True)
            p.restoreState()
        except Exception as e:
            print("Error al agregar la marca de agua:", e)

    # Añadir títulos
    p.setFont("Helvetica-Bold", 24)
    title_text = "SAMSAMANA"
    title_width = p.stringWidth(title_text, "Helvetica-Bold", 24)
    p.drawString((width - title_width) / 2, y_position, title_text)
    y_position -= 30

    p.setFont("Helvetica", 18)
    subtitle_text = f"Detalle de Compra #{compra.id}"
    subtitle_width = p.stringWidth(subtitle_text, "Helvetica", 18)
    p.drawString((width - subtitle_width) / 2, y_position, subtitle_text)
    y_position -= 50

    # Información de la compra
    info_data = [
        ['Usuario', compra.usuario.username],
        ['Fecha', compra.fecha.strftime('%Y-%m-%d %H:%M:%S')],
        ['Valor Total', f"${compra.total_compra:.2f}"],
    ]

    # Añadir tabla de información de la compra
    column_widths = [100, 200]
    p.setFont("Helvetica-Bold", 12)
    for label, value in info_data:
        p.drawString(margin, y_position, label)
        p.drawString(margin + column_widths[0], y_position, value)
        y_position -= 20  # Espaciado entre filas

    # Detalles de la compra
    y_position -= 20  # Espacio antes de la tabla de detalles
    p.setFont("Helvetica-Bold", 12)
    p.drawString(margin, y_position, "Detalles de la Compra")
    y_position -= 20

    # Encabezados y datos de la tabla de detalles
    headers = ["Producto", "Proveedor", "Cantidad", "Precio", "Subtotal"]
    table_data = [headers]

    for detalle in compra.detalles.all():
        table_data.append([
            detalle.producto.nombre,
            detalle.producto.proveedor.nombre,  # Suponiendo que tienes un campo proveedor en el producto
            str(detalle.cantidad),
            f"${detalle.producto.precio:.2f}",
            f"${detalle.subtotal:.2f}"
        ])

    # Crear y estilizar la tabla
    table = Table(table_data, colWidths=[150, 150, 100, 100, 100])  # Ajusta el tamaño de las columnas

    style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#003366')),  # Encabezado azul oscuro
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),  # Texto blanco en encabezados
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),  # Centrar texto
        ('GRID', (0, 0), (-1, -1), 1, colors.black),  # Bordes negros
        ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#f2f2f2')),  # Fondo gris claro para datos
        ('FONT', (0, 0), (-1, 0), 'Helvetica-Bold', 12),  # Fuente en encabezados
        ('FONT', (0, 1), (-1, -1), 'Helvetica', 10),  # Fuente en datos
        ('INNERGRID', (0, 0), (-1, -1), 0.5, colors.black),  # Bordes internos más delgados
        ('BOX', (0, 0), (-1, -1), 1, colors.black),  # Borde externo
    ])

    table.setStyle(style)

    # Ajustar la posición de la tabla
    table_width, table_height = table.wrap(width - 2 * margin, height)
    table_x = (width - table_width) / 2  # Centrar la tabla
    table_y = y_position - table_height - 20  # Espaciado adicional

    table.drawOn(p, table_x, table_y)

    # Finalizar el PDF
    p.showPage()
    p.save()
    buffer.seek(0)

    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename=detalle_compra_{compra.id}.pdf'
    return response


def exportar_detalle_compra_excel(request, id):
    compra = get_object_or_404(Compra, id=id)
    buffer = BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = f"Detalle Compra {compra.id}"

    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="0066cc", end_color="0066cc", fill_type="solid")
    alignment_center = Alignment(horizontal="center", vertical="center")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # Agregar logo
    logo_path = os.path.join(settings.STATIC_ROOT, 'img', 'Samsamanalogo1PNG.png')
    if os.path.exists(logo_path):
        img = Image(logo_path)
        img.width = 170
        img.height = 80
        ws.add_image(img, 'A1')

    # Título
    ws.merge_cells('A2:F2')
    title_cell = ws['A2']
    title_cell.value = f"DETALLE DE COMPRA #{compra.id} - BALNEARIO SAMSAMANA"
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = alignment_center

    # Información de la compra
    info_data = [
        ['Usuario', compra.usuario.username],
        ['Fecha', compra.fecha.strftime('%Y-%m-%d %H:%M:%S')],
        ['Valor Total', f"${compra.total_compra:.2f}"],
    ]
    for row, (label, value) in enumerate(info_data, start=4):
        ws.cell(row=row, column=1, value=label).alignment = alignment_center
        ws.cell(row=row, column=2, value=value).alignment = alignment_center

    # Encabezados de los detalles
    headers = ["Producto", "Proveedor", "Cantidad", "Precio", "Subtotal"]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=7, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = alignment_center

    # Detalles de la compra
    row_num = 8
    for detalle in compra.detalles.all():
        ws.cell(row=row_num, column=1, value=detalle.producto.nombre).alignment = alignment_center
        ws.cell(row=row_num, column=2, value=detalle.producto.proveedor.nombre).alignment = alignment_center  # Proveedor
        ws.cell(row=row_num, column=3, value=detalle.cantidad).alignment = alignment_center
        ws.cell(row=row_num, column=4, value=f"${detalle.producto.precio:.2f}").alignment = alignment_center
        ws.cell(row=row_num, column=5, value=f"${detalle.subtotal:.2f}").alignment = alignment_center
        row_num += 1

    # Aplicar bordes y centrar contenido
    for row in ws[f'A4:F{row_num - 1}']:
        for cell in row:
            cell.border = thin_border
            cell.alignment = alignment_center

    # Ajustar anchos de columna
    column_widths = [30, 30, 15, 15, 15]  # Ajusta el ancho de la columna del proveedor
    for col_num, width in enumerate(column_widths, start=1):
        ws.column_dimensions[chr(64 + col_num)].width = width

    # Guardar archivo Excel
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(buffer.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=detalle_compra_{compra.id}.xlsx'
    return response