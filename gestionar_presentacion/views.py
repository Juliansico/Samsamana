from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from .models import Presentacion
from .forms import PresentacionForm
from django.views.decorators.cache import never_cache
from io import BytesIO
from django.http import HttpResponse
from django.conf import settings
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from reportlab.lib.utils import ImageReader
from reportlab.platypus import Table, TableStyle
from reportlab.lib import colors
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image
from reportlab.lib import colors
from openpyxl.styles import Border,Side

def get_breadcrumbs(request):
    path = request.path.split('/')[1:]
    breadcrumbs = [{'title': 'Inicio', 'url': '/dashboard/'}]  # El inicio te lleva al dashboard
    url = ''
    for item in path:
        url += f'/{item}'
        breadcrumbs.append({'title': item.capitalize(), 'url': url})
    return breadcrumbs

@never_cache
def dashboard(request):
    return render(request, 'dashboard.html')


@never_cache
@login_required

def gestionar_presentacion(request):
    presentaciones = Presentacion.objects.all()
    if request.method == 'POST':
        form = PresentacionForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Presentación creada con éxito.')
            return redirect('gestionar_presentacion')
    else:
        form = PresentacionForm()
    breadcrumbs = get_breadcrumbs(request)
    return render(request, 'gestionar_presentacion.html', {'presentaciones': presentaciones, 'form': form, 'breadcrumbs': breadcrumbs})


@never_cache
@login_required
def añadir_presentacion(request):
    if request.method == 'POST':
        form = PresentacionForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Presentación añadida con éxito.')
            return redirect('gestionar_presentacion')
    else:
        form = PresentacionForm()
    breadcrumbs = get_breadcrumbs(request)
    return render(request, 'añadir_presentacion.html', {'form': form, 'breadcrumbs': breadcrumbs})

@never_cache
@login_required
def editar_presentacion(request, presentacion_id):
    presentacion = get_object_or_404(Presentacion, id=presentacion_id)
    if request.method == 'POST':
        form = PresentacionForm(request.POST, instance=presentacion)
        if form.is_valid():
            form.save()
            messages.success(request, 'Presentación actualizada con éxito.')
            return redirect('gestionar_presentacion')
    else:
        form = PresentacionForm(instance=presentacion)
    breadcrumbs = get_breadcrumbs(request)
    return render(request, 'editar_presentacion.html', {'form': form, 'presentacion': presentacion, 'breadcrumbs': breadcrumbs})


@never_cache
@login_required
def activar_inactivar_presentacion(request, presentacion_id):
    presentacion = get_object_or_404(Presentacion, id=presentacion_id)
    presentacion.estado = not presentacion.estado
    presentacion.save()
    estado = "activada" if presentacion.estado else "inactivada"
    messages.success(request, f'Presentación {estado} con éxito.')
    breadcrumbs = get_breadcrumbs(request)
    return redirect('gestionar_presentacion')


@never_cache
@login_required
def filtrar_presentaciones(request):
    estado_filtro = request.GET.get('estado', None)
    buscar = request.GET.get('buscar', '')

    presentaciones = Presentacion.objects.all()

    if estado_filtro == 'activado':
        presentaciones = presentaciones.filter(estado=True)
    elif estado_filtro == 'inactivado':
        presentaciones = presentaciones.filter(estado=False)

    if buscar:
        presentaciones = presentaciones.filter(nombre__icontains=buscar)

    breadcrumbs = get_breadcrumbs(request)
    context = {
        'presentaciones': presentaciones,
        'breadcrumbs': breadcrumbs
    }

    return render(request, 'gestionar_presentacion.html', context)

def reporte_presentaciones_pdf(request):
    buffer = BytesIO()
    p = canvas.Canvas(buffer, pagesize=letter)
    width, height = letter
    margin = 50
    row_height = 20
    y_position = height - margin

    # Cargar la imagen de marca de agua
    watermark_path = os.path.join(settings.STATIC_ROOT, 'img', 'Samsamanalogo1PNG.png')
    if os.path.exists(watermark_path):
        try:
            p.saveState()
            p.setFillAlpha(0.1)  # Ajusta la transparencia (0.1 = 10% visible)
            img = ImageReader(watermark_path)
            img_width, img_height = img.getSize()
            aspect = img_height / float(img_width)
            p.drawImage(img, x=(width - img_width * aspect) / 2, y=(height - img_height * aspect) / 2, 
                        width=img_width * aspect, height=img_height * aspect, mask='auto', preserveAspectRatio=True)
            p.restoreState()
        except Exception as e:
            print("Error al agregar la marca de agua:", e)

    # Añadir títulos
    p.setFont("Helvetica-Bold", 24)
    title_text = "SAMSAMANA"
    title_width = p.stringWidth(title_text, "Helvetica-Bold", 24)
    p.drawString((width - title_width) / 2, y_position, title_text)
    y_position -= 30

    p.setFont("Helvetica", 18)
    subtitle_text = "Reporte de Presentaciones"
    subtitle_width = p.stringWidth(subtitle_text, "Helvetica", 18)
    p.drawString((width - subtitle_width) / 2, y_position, subtitle_text)
    y_position -= 50

    # Crear tabla con datos de presentaciones
    column_widths = [50, 150, 60]
    headers = ["ID", "Nombre", "Estado"]
    data = [headers]

    presentaciones = Presentacion.objects.all()
    for presentacion in presentaciones:
        data.append([
            str(presentacion.id),
            presentacion.nombre,
            'Activo' if presentacion.estado else 'Inactivo'
        ])

    table = Table(data, colWidths=column_widths, rowHeights=row_height)

    # Estilo de la tabla
    style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#003366')),  # Fondo oscuro para encabezados
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),  # Texto blanco en encabezados
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),  # Bordes negros
        ('FONT', (0, 0), (-1, 0), 'Helvetica-Bold', 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#f2f2f2')),  # Fondo gris claro para datos
        ('FONT', (0, 1), (-1, -1), 'Helvetica', 10),
        ('INNERGRID', (0, 0), (-1, -1), 0.5, colors.black),  # Bordes internos más delgados
        ('BOX', (0, 0), (-1, -1), 1, colors.black),  # Borde externo
    ])
    table.setStyle(style)

    # Ajustar la posición de la tabla
    table_width, table_height = table.wrap(width, height)
    table_x = (width - table_width) / 2
    table_y = y_position - len(data) * row_height

    table.drawOn(p, table_x, table_y)

    p.showPage()
    p.save()
    buffer.seek(0)

    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="Reporte_presentaciones.pdf"'
    return response


def reporte_presentaciones_excel(request):
    buffer = BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte de Presentaciones"

    # Estilos
    header_font = Font(bold=True, color="FFFFFF")  # Fuente blanca y en negrita para encabezados
    header_fill = PatternFill(start_color="0066cc", end_color="0066cc", fill_type="solid")  # Fondo azul oscuro para encabezados
    alignment_center = Alignment(horizontal="center", vertical="center")  # Alineación centrada

    # Colores para el estado
    active_font = Font(color="00FF00")  # Verde para Activo
    inactive_font = Font(color="FF0000")  # Rojo para Inactivo

    # Añadir logo
    logo_path = os.path.join(settings.STATIC_ROOT, 'img', 'Samsamanalogo1PNG.png')
    if os.path.exists(logo_path):
        img = Image(logo_path)
        img.width = 170  # Ajustar tamaño del logo
        img.height = 100
        # Colocar el logo en A1
        ws.add_image(img, 'A1')

        # Ajustar tamaño de las celdas para centrar el logo visualmente
        ws.column_dimensions['A'].width = 20  # Ajustar el ancho de la columna A
        ws.column_dimensions['B'].width = 20  # Ajustar el ancho de la columna B

    # Añadir título
    ws.merge_cells('A4:C4')  # Combina celdas para centrar el título
    title_cell = ws['A4']
    title_cell.value = "PRESENTACIONES - BALNEARIO SAMSAMANA"
    title_cell.font = Font(bold=True, size=12)  # Fuente en negrita y tamaño 12 para el título
    title_cell.alignment = alignment_center  # Alineación centrada

    # Encabezados
    headers = ["ID", "Nombre", "Estado"]
    for col_num, header in enumerate(headers, 1):  # Cambiar para que los encabezados comiencen desde la columna A
        cell = ws.cell(row=6, column=col_num, value=header)
        cell.font = header_font  # Aplicar fuente blanca en negrita
        cell.fill = header_fill  # Aplicar fondo azul oscuro
        cell.alignment = alignment_center  # Alineación centrada

    # Datos
    presentaciones = Presentacion.objects.all()
    for row_num, presentacion in enumerate(presentaciones, 7):  # Los datos empiezan desde la fila 7
        ws.cell(row=row_num, column=1, value=presentacion.id).alignment = alignment_center  # Columna A para el ID
        ws.cell(row=row_num, column=2, value=presentacion.nombre).alignment = alignment_center  # Columna B para el nombre
        estado_cell = ws.cell(row=row_num, column=3, value='Activo' if presentacion.estado else 'Inactivo')  # Columna C para el estado
        estado_cell.alignment = alignment_center
        estado_cell.font = active_font if presentacion.estado else inactive_font  # Aplicar color verde o rojo según el estado

    # Ajustar anchos de columna
    column_widths = [20, 30, 15]  # Ajustar el ancho de columna según sea necesario
    for col_num, width in enumerate(column_widths, 1):  # A partir de la columna 1 (A)
        column_letter = get_column_letter(col_num)
        ws.column_dimensions[column_letter].width = width

    # Añadir bordes a las celdas de la tabla
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    for row in ws[f'A6:C{ws.max_row}']:  # Cambiar para aplicar bordes de A a C
        for cell in row:
            cell.border = thin_border

    # Ajustar altura de filas
    ws.row_dimensions[2].height = 70  # Ajustar altura del logo
    ws.row_dimensions[4].height = 25  # Altura del título
    ws.row_dimensions[6].height = 20  # Altura de los encabezados

    # Guardar y devolver
    wb.save(buffer)
    buffer.seek(0)
    response = HttpResponse(buffer.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="Reporte_presentaciones_samsamana.xlsx"'
    return response