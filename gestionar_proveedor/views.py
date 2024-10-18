from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from .models import Proveedor
from .forms import ProveedorForm
from django.views.decorators.cache import never_cache
from io import BytesIO
from openpyxl.drawing.image import Image
from django.http import HttpResponse
from django.conf import settings
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from reportlab.lib.utils import ImageReader
from reportlab.platypus import Table, TableStyle
from reportlab.lib import colors
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.styles import Border, Side, Alignment, Font, PatternFill
from reportlab.lib import colors
import os


def get_breadcrumbs(request):
    path = request.path.split('/')[1:]
    breadcrumbs = [{'title': 'Inicio', 'url': '/dashboard/'}]  # El inicio te lleva al dashboard
    url = ''
    for item in path:
        url += f'/{item}'
        breadcrumbs.append({'title': item.capitalize(), 'url': url})
    return breadcrumbs


@never_cache
def dashboard(request):
    return render(request, 'dashboard.html')
@never_cache
@login_required
def gestionar_proveedor(request):
    proveedores = Proveedor.objects.all()
    breadcrumbs = get_breadcrumbs(request)
    return render(request, 'gestionar_proveedor.html', {'proveedores': proveedores, 'breadcrumbs': breadcrumbs})
@never_cache
@login_required
def añadir_proveedor(request):
    if request.method == 'POST':
        form = ProveedorForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Proveedor añadido con éxito.')
            return redirect('gestionar_proveedor')
    else:
        form = ProveedorForm()
    breadcrumbs = get_breadcrumbs(request)
    return render(request, 'añadir_proveedor.html', {'form': form, 'breadcrumbs': breadcrumbs})


@never_cache
@login_required
def editar_proveedor(request, proveedor_id):
    proveedor = get_object_or_404(Proveedor, id=proveedor_id)
    if request.method == 'POST':
        form = ProveedorForm(request.POST, instance=proveedor)
        if form.is_valid():
            form.save()
            messages.success(request, 'Proveedor actualizado con éxito.')
            return redirect('gestionar_proveedor')
    else:
        form = ProveedorForm(instance=proveedor)
    breadcrumbs = get_breadcrumbs(request)
    return render(request, 'editar_proveedor.html', {'form': form, 'proveedor': proveedor, 'breadcrumbs': breadcrumbs})


@never_cache
@login_required
def activar_inactivar_proveedor(request, proveedor_id):
    proveedor = get_object_or_404(Proveedor, id=proveedor_id)
    proveedor.estado = not proveedor.estado
    proveedor.save()
    estado = "activado" if proveedor.estado else "inactivado"
    messages.success(request, f'Proveedor {estado} con éxito.')
    breadcrumbs = get_breadcrumbs(request)
    return redirect('gestionar_proveedor')


@never_cache
@login_required
def filtrar_proveedores(request):
    nombre_filtro = request.GET.get('nombre', '')
    estado_filtro = request.GET.get('estado', None)
    telefono_filtro = request.GET.get('telefono', '')
    email_filtro = request.GET.get('email', '')

    proveedores = Proveedor.objects.all()

    if nombre_filtro:
        proveedores = proveedores.filter(nombre__icontains=nombre_filtro)
    if estado_filtro == 'activado':
        proveedores = proveedores.filter(estado=True)
    elif estado_filtro == 'inactivado':
        proveedores = proveedores.filter(estado=False)
    if telefono_filtro:
        proveedores = proveedores.filter(telefono__icontains=telefono_filtro)
    if email_filtro:
        proveedores = proveedores.filter(email__icontains=email_filtro)

    breadcrumbs = get_breadcrumbs(request)
    context = {
        'proveedores': proveedores,
        'breadcrumbs': breadcrumbs
    }

    return render(request, 'gestionar_proveedor.html', context)

def reporte_proveedores_pdf(request):
    buffer = BytesIO()
    p = canvas.Canvas(buffer, pagesize=letter)
    width, height = letter
    margin = 50
    row_height = 20
    y_position = height - margin

    # Cargar la imagen de marca de agua
    watermark_path = os.path.join(settings.STATIC_ROOT, 'img', 'Samsamanalogo1PNG.png')
    if os.path.exists(watermark_path):
        try:
            p.saveState()
            p.setFillAlpha(0.1)  # Ajusta la transparencia al 10%
            img = ImageReader(watermark_path)
            img_width, img_height = img.getSize()
            aspect = img_height / float(img_width)
            p.drawImage(img, x=(width - img_width * aspect) / 2, y=(height - img_height * aspect) / 2, 
                        width=img_width * aspect, height=img_height * aspect, mask='auto', preserveAspectRatio=True)
            p.restoreState()
        except Exception as e:
            print("Error al agregar la marca de agua:", e)

    # Añadir títulos
    p.setFont("Helvetica-Bold", 24)
    title_text = "SAMSAMANA"
    title_width = p.stringWidth(title_text, "Helvetica-Bold", 24)
    p.drawString((width - title_width) / 2, y_position, title_text)
    y_position -= 30

    p.setFont("Helvetica", 18)
    subtitle_text = "Reporte de Proveedores"
    subtitle_width = p.stringWidth(subtitle_text, "Helvetica", 18)
    p.drawString((width - subtitle_width) / 2, y_position, subtitle_text)
    y_position -= 50

    # Crear tabla con datos de proveedores
    column_widths = [50, 150, 100, 200, 100]
    headers = ["ID", "Nombre", "Teléfono", "Email", "Estado"]
    data = [headers]

    proveedores = Proveedor.objects.all()
    for proveedor in proveedores:
        data.append([
            str(proveedor.id),
            proveedor.nombre,
            proveedor.telefono,
            proveedor.email,
            'Activo' if proveedor.estado else 'Inactivo'
        ])

    table = Table(data, colWidths=column_widths, rowHeights=row_height)

    # Estilo de la tabla
    style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#003366')),  # Fondo oscuro para encabezados
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),  # Texto blanco en encabezados
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),  # Bordes negros
        ('FONT', (0, 0), (-1, 0), 'Helvetica-Bold', 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#f2f2f2')),  # Fondo gris claro para datos
        ('FONT', (0, 1), (-1, -1), 'Helvetica', 10),
        ('INNERGRID', (0, 0), (-1, -1), 0.5, colors.black),  # Bordes internos más delgados
        ('BOX', (0, 0), (-1, -1), 1, colors.black),  # Borde externo
    ])
    table.setStyle(style)

    # Ajustar la posición de la tabla
    table_width, table_height = table.wrap(width, height)
    table_x = (width - table_width) / 2
    table_y = y_position - len(data) * row_height

    table.drawOn(p, table_x, table_y)

    p.showPage()
    p.save()
    buffer.seek(0)

    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="Reporte_proveedores.pdf"'
    return response

def reporte_proveedores_excel(request):
    buffer = BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte de Proveedores"

    # Estilos
    header_font = Font(bold=True, color="FFFFFF")  
    header_fill = PatternFill(start_color="0066cc", end_color="0066cc", fill_type="solid")  
    alignment_center = Alignment(horizontal="center", vertical="center") 

    # Colores para el estado
    active_font = Font(color="00FF00")  
    inactive_font = Font(color="FF0000")  

    # Añadir logo
    logo_path = os.path.join(settings.STATIC_ROOT, 'img', 'Samsamanalogo1PNG.png')
    if os.path.exists(logo_path):
        img = Image(logo_path)
        img.width = 170  
        img.height = 100
        ws.add_image(img, 'A2')

        ws.column_dimensions['B'].width = 15  
        ws.column_dimensions['C'].width = 15 
        ws.column_dimensions['D'].width = 15  
        ws.column_dimensions['E'].width = 15  
        ws.column_dimensions['F'].width = 15  

    # Añadir título
    ws.merge_cells('B4:F4')  
    title_cell = ws['B4']
    title_cell.value = "TABLA PROVEEDORES - BALNEARIO SAMSAMANA"
    title_cell.font = Font(bold=True, size=16)  
    title_cell.alignment = alignment_center  

    # Encabezados
    headers = ["ID", "Nombre", "Teléfono", "Email", "Estado"]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=6, column=col_num, value=header)
        cell.font = header_font  
        cell.fill = header_fill  
        cell.alignment = alignment_center  

    # Datos
    proveedores = Proveedor.objects.all()
    for row_num, proveedor in enumerate(proveedores, 7):  
        ws.cell(row=row_num, column=1, value=proveedor.id).alignment = alignment_center
        ws.cell(row=row_num, column=2, value=proveedor.nombre).alignment = alignment_center
        ws.cell(row=row_num, column=3, value=proveedor.telefono).alignment = alignment_center
        ws.cell(row=row_num, column=4, value=proveedor.email).alignment = alignment_center
        estado_cell = ws.cell(row=row_num, column=5, value='Activo' if proveedor.estado else 'Inactivo')
        estado_cell.alignment = alignment_center
        estado_cell.font = active_font if proveedor.estado else inactive_font  

    
    column_widths = [20, 30, 20, 30, 15]  
    for col_num, width in enumerate(column_widths, 1):
        column_letter = get_column_letter(col_num)
        ws.column_dimensions[column_letter].width = width

    
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    for row in ws[f'B6:F{ws.max_row}']:
        for cell in row:
            cell.border = thin_border

    
    ws.row_dimensions[2].height = 40  
    ws.row_dimensions[4].height = 25  
    ws.row_dimensions[6].height = 20  

    
    wb.save(buffer)
    buffer.seek(0)
    response = HttpResponse(buffer.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="Reporte_proveedores_samsamana.xlsx"'
    return response