from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required 
from django.contrib import messages
from .models import Marca
from .forms import MarcaForm
from django.views.decorators.cache import never_cache
from io import BytesIO
from django.http import HttpResponse
from django.conf import settings
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from reportlab.lib.utils import ImageReader
from reportlab.platypus import Table, TableStyle
from reportlab.lib import colors
from .models import Marca
import os
from openpyxl.utils import get_column_letter
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.drawing.image import Image
from reportlab.lib import colors
from openpyxl.styles import Border,Side

def get_breadcrumbs(request):
    path = request.path.split('/')[1:]
    breadcrumbs = [{'title': 'Inicio', 'url': '/dashboard/'}]  # El inicio te lleva al dashboard
    url = ''
    for item in path:
        url += f'/{item}'
        breadcrumbs.append({'title': item.capitalize(), 'url': url})
    return breadcrumbs

@never_cache
def dashboard(request): 
    return render(request, 'dashboard.html')
@never_cache
@login_required
def editar_marca(request, marca_id):
    # Buscar la marca usando filter y first en lugar de get_object_or_404
    marca = Marca.objects.filter(id=marca_id).first()
    
    if marca is None:
        # Si la marca no se encuentra, mostrar un mensaje de error y redirigir
        messages.error(request, 'Marca no encontrada.')
        return redirect('gestionar_marca')

    if request.method == 'POST':
        form = MarcaForm(request.POST, request.FILES, instance=marca)
        if form.is_valid():
            form.save()
            messages.success(request, 'Marca actualizada con éxito.')
            return redirect('gestionar_marca')
    else:
        form = MarcaForm(instance=marca)
    
    breadcrumbs = get_breadcrumbs(request)
    return render(request, 'editar_marca.html', {'form': form, 'marca': marca, 'breadcrumbs': breadcrumbs})
@never_cache
@login_required
def activar_inactivar_marca(request, marca_id):
    # Buscar la marca usando filter y first en lugar de get_object_or_404
    marca = Marca.objects.filter(id=marca_id).first()
    
    if marca is None:
        # Si la marca no se encuentra, mostrar un mensaje de error y redirigir
        messages.error(request, 'Marca no encontrada.')
        return redirect('gestionar_marca')

    # Cambiar el estado de la marca
    marca.estado = not marca.estado
    marca.save()
    estado = "activada" if marca.estado else "inactivada"
    messages.success(request, f'Marca {estado} con éxito.')
    breadcrumbs = get_breadcrumbs(request)
    return redirect('gestionar_marca')
@never_cache
@login_required
def filtrar_marcas(request):
    nombre_filtro = request.GET.get('buscar', '')
    estado_filtro = request.GET.get('estado', '')

    marcas = Marca.objects.all()

    if nombre_filtro:
        marcas = marcas.filter(nombre__icontains=nombre_filtro)

    if estado_filtro == 'activado':
        marcas = marcas.filter(estado=True)
    elif estado_filtro == 'inactivado':
        marcas = marcas.filter(estado=False)

    breadcrumbs = get_breadcrumbs(request)
    context = {
        'marcas': marcas,
        'breadcrumbs': breadcrumbs,
    }

    return render(request, 'gestionar_marca.html', context)
@never_cache
@login_required
def gestionar_marca(request):
    marcas = Marca.objects.all()
    breadcrumbs = get_breadcrumbs(request)
    return render(request, 'gestionar_marca.html', {'marcas': marcas, 'breadcrumbs': breadcrumbs})
@never_cache
@login_required
def añadir_marca(request):
    if request.method == 'POST':
        form = MarcaForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            messages.success(request, 'Marca añadida con éxito.')
            return redirect('gestionar_marca')
    else:
        form = MarcaForm()
    breadcrumbs = get_breadcrumbs(request)
    return render(request, 'añadir_marca.html', {'form': form, 'breadcrumbs': breadcrumbs})

def reporte_marcas_pdf(request):
    buffer = BytesIO()
    p = canvas.Canvas(buffer, pagesize=letter)
    width, height = letter
    margin = 50
    row_height = 20
    y_position = height - margin

    # Cargar la imagen de marca de agua
    watermark_path = os.path.join(settings.STATIC_ROOT, 'img', 'Samsamanalogo1PNG.png')
    if os.path.exists(watermark_path):
        try:
            p.saveState()
            p.setFillAlpha(0.1)  # Ajusta la transparencia (0.1 = 10% visible)
            img = ImageReader(watermark_path)
            img_width, img_height = img.getSize()
            aspect = img_height / float(img_width)
            p.drawImage(img, x=(width - img_width * aspect) / 2, y=(height - img_height * aspect) / 2, 
                        width=img_width * aspect, height=img_height * aspect, mask='auto', preserveAspectRatio=True)
            p.restoreState()
        except Exception as e:
            print("Error al agregar la marca de agua:", e)

    # Añadir títulos
    p.setFont("Helvetica-Bold", 24)
    title_text = "SAMSAMANA"
    title_width = p.stringWidth(title_text, "Helvetica-Bold", 24)
    p.drawString((width - title_width) / 2, y_position, title_text)
    y_position -= 30

    p.setFont("Helvetica", 18)
    subtitle_text = "Reporte de Marcas"
    subtitle_width = p.stringWidth(subtitle_text, "Helvetica", 18)
    p.drawString((width - subtitle_width) / 2, y_position, subtitle_text)
    y_position -= 50

    # Crear tabla con datos de marcas
    column_widths = [50, 150, 60]
    headers = ["ID", "Nombre", "Estado"]
    data = [headers]

    marcas = Marca.objects.all()
    for marca in marcas:
        data.append([
            str(marca.id),
            marca.nombre,
            'Activo' if marca.estado else 'Inactivo'
        ])

    table = Table(data, colWidths=column_widths, rowHeights=row_height)

    # Estilo de la tabla
    style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#003366')),  # Fondo oscuro para encabezados
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),  # Texto blanco en encabezados
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),  # Bordes negros
        ('FONT', (0, 0), (-1, 0), 'Helvetica-Bold', 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#f2f2f2')),  # Fondo gris claro para datos
        ('FONT', (0, 1), (-1, -1), 'Helvetica', 10),
        ('INNERGRID', (0, 0), (-1, -1), 0.5, colors.black),  # Bordes internos más delgados
        ('BOX', (0, 0), (-1, -1), 1, colors.black),  # Borde externo
    ])
    table.setStyle(style)

    # Ajustar la posición de la tabla
    table_width, table_height = table.wrap(width, height)
    table_x = (width - table_width) / 2
    table_y = y_position - len(data) * row_height

    table.drawOn(p, table_x, table_y)

    p.showPage()
    p.save()
    buffer.seek(0)

    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="Reporte_marcas.pdf"'
    return response

def reporte_marcas_excel(request):
    # Crear buffer para el archivo Excel
    buffer = BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte de Marcas"

    # Estilos de Excel
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="0066cc", end_color="0066cc", fill_type="solid")
    alignment_center = Alignment(horizontal="center", vertical="center")

    # Colores para el estado
    active_font = Font(color="00FF00")  # Verde para Activo
    inactive_font = Font(color="FF0000")  # Rojo para Inactivo

    # Añadir logo
    logo_path = os.path.join(settings.STATIC_ROOT, 'img', 'Samsamanalogo1PNG.png')
    if os.path.exists(logo_path):
        img = Image(logo_path)
        img.width = 150
        img.height = 75
        ws.add_image(img, 'A1')  # Posicionar logo en A1

        # Ajustar tamaño de las celdas para centrar el logo visualmente
        ws.column_dimensions['A'].width = 25  # Ajustar el ancho de la columna A
        ws.row_dimensions[1].height = 100  # Ajustar la altura de la fila 1 para el logo

    # Añadir título
    ws.merge_cells('A4:C4')  # Combina celdas para centrar el título
    title_cell = ws['A4']
    title_cell.value = "TABLA MARCAS"
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = alignment_center

    # Encabezados de la tabla
    headers = ["ID", "Nombre", "Estado"]
    for col_num, header in enumerate(headers, 1):  # Encabezados empiezan en la columna A
        cell = ws.cell(row=6, column=col_num, value=header)  # Fila 6 para encabezados
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = alignment_center

    # Obtener datos de marcas del modelo y agregar a la tabla
    marcas = Marca.objects.all()
    for row_num, marca in enumerate(marcas, start=7):  # Los datos empiezan desde la fila 7
        ws.cell(row=row_num, column=1, value=marca.id).alignment = alignment_center
        ws.cell(row=row_num, column=2, value=marca.nombre).alignment = alignment_center
        estado = 'Activo' if marca.estado else 'Inactivo'
        estado_font = active_font if marca.estado else inactive_font
        estado_cell = ws.cell(row=row_num, column=3, value=estado)
        estado_cell.alignment = alignment_center
        estado_cell.font = estado_font

    # Ajustar anchos de columna
    column_widths = [20, 30, 15]  # Ajustar el ancho de columna según sea necesario
    for col_num, width in enumerate(column_widths, 1):  # Ajustar columnas A, B, C
        column_letter = get_column_letter(col_num)
        ws.column_dimensions[column_letter].width = width

    # Agregar bordes a la tabla
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    for row in ws[f'A6:C{ws.max_row}']:
        for cell in row:
            cell.border = thin_border

    # Ajustar alturas de filas
    ws.row_dimensions[4].height = 30  # Altura para la fila del título
    ws.row_dimensions[6].height = 20  # Altura para la fila de encabezados

    # Guardar y devolver el archivo Excel
    wb.save(buffer)
    buffer.seek(0)
    response = HttpResponse(buffer.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="Reporte_marcas_samsamana.xlsx"'
    return response